"""
Excel Manager for AutoBE

Handles Excel/CSV-based addon organisation system for modpack configurations.
When *openpyxl* is unavailable the manager transparently falls back to CSV
storage.

Exposed via the module-level :func:`is_excel_available` predicate.
"""

from __future__ import annotations

import csv as _csv_module
import json as _json_module
import logging as _logging_module
from datetime import datetime as datetime
from pathlib import Path
from typing import Any

try:
    from packaging import version as _pkg_version

    _PACKAGING_AVAILABLE = True
except ImportError:
    _PACKAGING_AVAILABLE = False

try:
    import openpyxl as _openpyxl
    from openpyxl import Workbook as _Workbook
    from openpyxl.styles import Font as _Font, PatternFill as _PatternFill, Alignment as _Alignment

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

_logger = _logging_module.getLogger(__name__)


# ── ExcelManager ───────────────────────────────────────────────────────────

class ExcelManager:
    """Manage modpack configurations via Excel (``.xlsx``) or CSV (``.csv``) files."""

    def __init__(self) -> None:
        self.workbook: Any = None
        self.current_file: str | None = None
        self.csv_mode: bool = not _OPENPYXL_AVAILABLE
        self.csv_data: dict[str, Any] = {}

    # ── Creation ────────────────────────────────────────────────────────────

    def create_new_workbook(self) -> None:
        """Initialise a new workbook (or CSV data structure)."""
        if self.csv_mode:
            self.csv_data = {}
            _logger.info("Created new CSV-based addon organization")
            return

        self.workbook = _Workbook()
        # Remove the default placeholder sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        self._build_summary_sheet()
        _logger.info("Created new Excel workbook for addon organization")

    def _build_summary_sheet(self) -> None:
        """Create the ``Modpack Summary`` sheet with styled headers."""
        sheet = self.workbook.create_sheet("Modpack Summary", 0)
        headers = ["Modpack Name", "Min Version", "Max Version", "Addon Count", "Last Updated"]
        sheet.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = _Font(bold=True)
            cell.fill = _PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = _Alignment(horizontal="center")

        sheet.freeze_panes = "A2"

    # ── Modpack sheet ─────────────────────────────────────────────────

    def add_modpack_sheet(
        self,
        modpack_name: str,
        addons: list[dict[str, Any]],
        min_version: str = "1.21.0",
        max_version: str = "1.21.90",
    ) -> None:
        """Add a new sheet (or CSV table) for a modpack and its addons."""
        if self.csv_mode:
            safe = self._sanitise_name(modpack_name)
            self.csv_data[safe] = {
                "name": modpack_name,
                "min_version": min_version,
                "max_version": max_version,
                "addons": self._sort_addons(addons),
            }
            _logger.info("Added modpack '%s' with %d addon(s) (CSV mode)", safe, len(addons))
            return

        if not self.workbook:
            self.create_new_workbook()

        safe = self._sanitise_name(modpack_name)

        if safe in self.workbook.sheetnames:
            _logger.warning("Sheet '%s' already exists — overwriting", safe)
            self.workbook.remove(self.workbook[safe])

        sheet = self.workbook.create_sheet(safe)
        sheet.append(["Modpack:", modpack_name])
        sheet.append(["Min Version:", min_version])
        sheet.append(["Max Version:", max_version])
        sheet.append([])

        for row in range(1, 4):
            sheet.cell(row=row, column=1).font = _Font(bold=True)

        headers = [
            "Addon Name", "File Path", "Addon Version",
            "Min Version", "Max Version", "Status", "Notes",
        ]
        sheet.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=5, column=col_idx)
            cell.font = _Font(bold=True)
            cell.fill = _PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = _Alignment(horizontal="center")

        for addon in self._sort_addons(addons):
            sheet.append([
                addon.get("name", ""),
                addon.get("path", ""),
                addon.get("version", ""),
                addon.get("min_version", ""),
                addon.get("max_version", ""),
                addon.get("status", "Unknown"),
                addon.get("notes", ""),
            ])

        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

        sheet.freeze_panes = "A6"
        self._refresh_summary(modpack_name, min_version, max_version, len(addons))
        _logger.info("Added modpack sheet '%s' with %d addon(s)", safe, len(addons))

    # ── Loading ───────────────────────────────────────────────────────

    def load_from_excel(self, file_path: str) -> dict[str, Any]:
        """Parse an Excel workbook (or CSV file) and return all modpack configs."""
        if self.csv_mode:
            return self._load_csv(file_path)

        self.workbook = _openpyxl.load_workbook(file_path)
        self.current_file = file_path

        configs: dict[str, Any] = {}
        for name in self.workbook.sheetnames:
            if name == "Modpack Summary":
                continue
            configs[name] = self._parse_sheet(self.workbook[name])

        _logger.info("Loaded %d modpack configuration(s) from %s", len(configs), file_path)
        return configs

    def _parse_sheet(self, sheet: Any) -> dict[str, Any]:
        """Extract a single modpack configuration from an openpyxl *sheet*."""
        config: dict[str, Any] = {
            "name": sheet.cell(row=1, column=2).value,
            "min_version": sheet.cell(row=2, column=2).value,
            "max_version": sheet.cell(row=3, column=2).value,
            "addons": [],
        }
        for row in range(6, sheet.max_row + 1):
            name = sheet.cell(row=row, column=1).value
            if not name:
                continue
            config["addons"].append({
                "name": name,
                "path": sheet.cell(row=row, column=2).value,
                "version": sheet.cell(row=row, column=3).value,
                "min_version": sheet.cell(row=row, column=4).value,
                "max_version": sheet.cell(row=row, column=5).value,
                "status": sheet.cell(row=row, column=6).value,
                "notes": sheet.cell(row=row, column=7).value,
            })
        return config

    # ── Saving ─────────────────────────────────────────────────────────

    def save_workbook(self, file_path: str) -> None:
        """Persist the current workbook (or CSV data) to *file_path*."""
        if self.csv_mode:
            self._save_csv(file_path)
            return

        if not self.workbook:
            raise ValueError("No workbook to save — create or load one first.")
        self.workbook.save(file_path)
        self.current_file = file_path
        _logger.info("Saved workbook to %s", file_path)

    # ── Export / import convenience ──────────────────────────────────

    def export_modpack_to_excel(
        self,
        modpack_name: str,
        source_packs: list[str],
        output_dir: str,
        min_version: str = "1.21.0",
        max_version: str = "1.21.90",
    ) -> str:
        """Export a modpack configuration derived from *source_packs* to an Excel file."""
        if not self.workbook:
            self.create_new_workbook()

        addons = [
            {
                "name": Path(p).stem,
                "path": p,
                "version": "Unknown",
                "min_version": min_version,
                "max_version": max_version,
                "status": "Active",
                "notes": "",
            }
            for p in source_packs
        ]
        self.add_modpack_sheet(modpack_name, addons, min_version, max_version)

        out = Path(output_dir) / f"{modpack_name}_config.xlsx"
        self.save_workbook(str(out))
        _logger.info("Exported modpack '%s' to %s", modpack_name, out)
        return str(out)

    def import_modpack_from_excel(self, file_path: str, modpack_name: str) -> dict[str, Any]:
        """Import a single modpack configuration from an existing Excel file."""
        configs = self.load_from_excel(file_path)
        safe = self._sanitise_name(modpack_name)
        if safe not in configs:
            raise ValueError(f"Modpack '{modpack_name}' not found in {file_path}")
        return configs[safe]

    # ── Queries ──────────────────────────────────────────────────────

    def get_available_modpacks(self) -> list[str]:
        """Return the names of all modpacks currently loaded."""
        if self.csv_mode:
            return list(self.csv_data.keys())
        if not self.workbook:
            return []
        return [n for n in self.workbook.sheetnames if n != "Modpack Summary"]

    def close(self) -> None:
        """Release the open workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.current_file = None

    # ── Internal ──────────────────────────────────────────────────────

    def _sanitise_name(self, name: str) -> str:
        """Strip characters forbidden in Excel sheet names; truncate to 31 chars."""
        for ch in ("\\", "/", "*", "[", "]", ":", "?"):
            name = name.replace(ch, "_")
        return name[:31]

    def _sort_addons(self, addons: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Sort addons by version descending."""
        def _key(a: dict[str, Any]) -> Any:
            v = a.get("version", "0.0.0")
            if _PACKAGING_AVAILABLE:
                try:
                    return _pkg_version.parse(v)
                except Exception:
                    return v
            return v
        return sorted(addons, key=_key, reverse=True)

    def _update_summary(self, name: str, min_v: str, max_v: str, count: int) -> None:
        """Add or update an entry in the ``Modpack Summary`` sheet."""
        sheet = self.workbook["Modpack Summary"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == name:
                sheet.cell(row=row, column=2).value = min_v
                sheet.cell(row=row, column=3).value = max_v
                sheet.cell(row=row, column=4).value = count
                sheet.cell(row=row, column=5).value = datetime.now().strftime("%Y-%m-%d")
                return
        sheet.append([name, min_v, max_v, count, datetime.now().strftime("%Y-%m-%d")])

    # ── CSV helpers ──────────────────────────────────────────────────

    def _load_csv(self, file_path: str) -> dict[str, Any]:
        """Parse a CSV file into modpack configurations."""
        self.current_file = file_path
        configs: dict[str, Any] = {}

        try:
            with open(file_path, "r", newline="", encoding="utf-8") as fh:
                reader = _csv_module.DictReader(fh)
                for row in reader:
                    mp = row.get("modpack_name", "")
                    if not mp:
                        continue
                    if mp not in configs:
                        configs[mp] = {
                            "name": mp,
                            "min_version": row.get("min_version", "1.21.0"),
                            "max_version": row.get("max_version", "1.21.90"),
                            "addons": [],
                        }
                    configs[mp]["addons"].append({
                        "name": row.get("addon_name", ""),
                        "path": row.get("path", ""),
                        "version": row.get("version", ""),
                        "min_version": row.get("min_version", ""),
                        "max_version": row.get("max_version", ""),
                        "status": row.get("status", "Unknown"),
                        "notes": row.get("notes", ""),
                    })
        except Exception as exc:
            _logger.error("Failed to load CSV: %s", exc)
            raise

        _logger.info("Loaded %d modpack(s) from CSV %s", len(configs), file_path)
        return configs

    def _save_csv(self, file_path: str) -> None:
        """Persist ``csv_data`` to a CSV file."""
        if not self.csv_data:
            _logger.warning("No data to save to CSV")
            return

        fieldnames = [
            "modpack_name", "addon_name", "path", "version",
            "min_version", "max_version", "status", "notes",
        ]
        try:
            with open(file_path, "w", newline="", encoding="utf-8") as fh:
                writer = _csv_module.DictWriter(fh, fieldnames=fieldnames)
                writer.writeheader()
                for mp_name, mp_config in self.csv_data.items():
                    for addon in mp_config["addons"]:
                        writer.writerow({
                            "modpack_name": mp_name,
                            "addon_name": addon.get("name", ""),
                            "path": addon.get("path", ""),
                            "version": addon.get("version", ""),
                            "min_version": addon.get("min_version", ""),
                            "max_version": addon.get("max_version", ""),
                            "status": addon.get("status", ""),
                            "notes": addon.get("notes", ""),
                        })
        except Exception as exc:
            _logger.error("Failed to save CSV: %s", exc)
            raise

        self.current_file = file_path
        _logger.info("Saved CSV to %s", file_path)


# ── Module-level predicate ─────────────────────────────────────────────────

def is_excel_available() -> bool:
    """Always ``True`` — CSV fallback works even without *openpyxl*."""
    return True