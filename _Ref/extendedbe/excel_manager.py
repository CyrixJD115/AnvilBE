"""
Excel Manager for AutoBE
Handles Excel/CSV-based addon organization system for modpack configurations.
Uses CSV as fallback when openpyxl is not available.
"""

import json
import logging
import csv
from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    from packaging import version as pkg_version
    PACKAGING_AVAILABLE = True
except ImportError:
    PACKAGING_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

_logger = logging.getLogger(__name__)


class ExcelManager:
    """
    Manages Excel/CSV files for organizing addons and modpack configurations.
    Each sheet/CSV represents a modpack with its addons, versions, and compatibility info.
    Uses CSV as fallback when openpyxl is not available.
    """
    
    def __init__(self):
        self.workbook = None
        self.current_file = None
        self.csv_mode = not OPENPYXL_AVAILABLE  # Use CSV mode if openpyxl not available
        self.csv_data = {}  # Store CSV data when in CSV mode
        
    def create_new_workbook(self):
        """Create a new Excel workbook for addon organization."""
        if self.csv_mode:
            # CSV mode: initialize empty data structure
            self.csv_data = {}
            _logger.info("Created new CSV-based addon organization")
            return
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel functionality. Install with: pip install openpyxl")
        
        self.workbook = Workbook()
        # Remove default sheet
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # Create summary sheet
        self._create_summary_sheet()
        
        _logger.info("Created new Excel workbook for addon organization")
    
    def _create_summary_sheet(self):
        """Create a summary sheet listing all modpacks."""
        sheet = self.workbook.create_sheet("Modpack Summary", 0)
        
        # Headers
        headers = ["Modpack Name", "Min Version", "Max Version", "Addon Count", "Last Updated"]
        sheet.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Freeze header row
        sheet.freeze_panes = "A2"
    
    def add_modpack_sheet(
        self,
        modpack_name: str,
        addons: List[Dict[str, Any]],
        min_version: str = "1.21.0",
        max_version: str = "1.21.90"
    ):
        """
        Add a new sheet for a modpack with its addons.
        
        Args:
            modpack_name: Name of the modpack
            addons: List of addon dictionaries with keys: name, path, version, compatibility
            min_version: Minimum Minecraft version for the modpack
            max_version: Maximum Minecraft version for the modpack
        """
        if self.csv_mode:
            # CSV mode: store data in dictionary
            sheet_name = self._sanitize_sheet_name(modpack_name)
            sorted_addons = self._sort_addons_by_version(addons)
            self.csv_data[sheet_name] = {
                "name": modpack_name,
                "min_version": min_version,
                "max_version": max_version,
                "addons": sorted_addons
            }
            _logger.info(f"Added modpack '{sheet_name}' with {len(addons)} addons (CSV mode)")
            return
        
        if not self.workbook:
            self.create_new_workbook()
        
        # Sanitize sheet name (Excel has restrictions)
        sheet_name = self._sanitize_sheet_name(modpack_name)
        
        # Check if sheet already exists
        if sheet_name in self.workbook.sheetnames:
            _logger.warning(f"Sheet '{sheet_name}' already exists, overwriting")
            self.workbook.remove(self.workbook[sheet_name])
        
        sheet = self.workbook.create_sheet(sheet_name)
        
        # Add modpack info at top
        sheet.append(["Modpack:", modpack_name])
        sheet.append(["Min Version:", min_version])
        sheet.append(["Max Version:", max_version])
        sheet.append([])  # Empty row
        
        # Style modpack info
        for row in range(1, 4):
            sheet.cell(row=row, column=1).font = Font(bold=True)
        
        # Addon headers
        headers = ["Addon Name", "File Path", "Addon Version", "Min Version", "Max Version", "Status", "Notes"]
        sheet.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=5, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Sort addons by version (automatic version sorting)
        sorted_addons = self._sort_addons_by_version(addons)
        
        # Add addon data
        for addon in sorted_addons:
            row_data = [
                addon.get("name", ""),
                addon.get("path", ""),
                addon.get("version", ""),
                addon.get("min_version", ""),
                addon.get("max_version", ""),
                addon.get("status", "Unknown"),
                addon.get("notes", "")
            ]
            sheet.append(row_data)
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        sheet.freeze_panes = "A6"
        
        # Update summary sheet
        self._update_summary_sheet(modpack_name, min_version, max_version, len(addons))
        
        _logger.info(f"Added modpack sheet '{sheet_name}' with {len(addons)} addons")
    
    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name to comply with Excel restrictions."""
        # Excel sheet names: max 31 chars, no special chars
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        return sanitized[:31]
    
    def _sort_addons_by_version(self, addons: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Sort addons by version using semantic versioning.
        
        Args:
            addons: List of addon dictionaries
            
        Returns:
            Sorted list of addons by version (descending)
        """
        def version_key(addon):
            ver = addon.get("version", "0.0.0")
            if PACKAGING_AVAILABLE:
                try:
                    return pkg_version.parse(ver)
                except:
                    # If version parsing fails, use the string as fallback
                    return ver
            else:
                # Simple string comparison if packaging not available
                return ver
        
        return sorted(addons, key=version_key, reverse=True)
    
    def _update_summary_sheet(self, modpack_name: str, min_version: str, max_version: str, addon_count: int):
        """Update the summary sheet with modpack info."""
        sheet = self.workbook["Modpack Summary"]
        
        # Check if modpack already in summary
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == modpack_name:
                # Update existing entry
                sheet.cell(row=row, column=2).value = min_version
                sheet.cell(row=row, column=3).value = max_version
                sheet.cell(row=row, column=4).value = addon_count
                from datetime import datetime
                sheet.cell(row=row, column=5).value = datetime.now().strftime("%Y-%m-%d")
                return
        
        # Add new entry
        from datetime import datetime
        sheet.append([
            modpack_name,
            min_version,
            max_version,
            addon_count,
            datetime.now().strftime("%Y-%m-%d")
        ])
    
    def load_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Load modpack configuration from Excel/CSV file.
        
        Args:
            file_path: Path to Excel/CSV file
            
        Returns:
            Dictionary with modpack configurations
        """
        if self.csv_mode:
            # CSV mode: load from CSV file
            return self._load_from_csv(file_path)
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel functionality. Install with: pip install openpyxl")
        
        self.workbook = openpyxl.load_workbook(file_path)
        self.current_file = file_path
        
        configurations = {}
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name == "Modpack Summary":
                continue
            
            sheet = self.workbook[sheet_name]
            modpack_config = self._parse_modpack_sheet(sheet)
            configurations[sheet_name] = modpack_config
        
        _logger.info(f"Loaded {len(configurations)} modpack configurations from {file_path}")
        return configurations
    
    def _parse_modpack_sheet(self, sheet) -> Dict[str, Any]:
        """Parse a modpack sheet and extract configuration."""
        config = {
            "name": sheet.cell(row=1, column=2).value,
            "min_version": sheet.cell(row=2, column=2).value,
            "max_version": sheet.cell(row=3, column=2).value,
            "addons": []
        }
        
        # Start from row 6 (after headers)
        for row in range(6, sheet.max_row + 1):
            addon_name = sheet.cell(row=row, column=1).value
            if not addon_name:
                continue
            
            addon = {
                "name": addon_name,
                "path": sheet.cell(row=row, column=2).value,
                "version": sheet.cell(row=row, column=3).value,
                "min_version": sheet.cell(row=row, column=4).value,
                "max_version": sheet.cell(row=row, column=5).value,
                "status": sheet.cell(row=row, column=6).value,
                "notes": sheet.cell(row=row, column=7).value
            }
            config["addons"].append(addon)
        
        return config
    
    def save_workbook(self, file_path: str):
        """Save the workbook/CSV to a file."""
        if self.csv_mode:
            # CSV mode: save as CSV
            return self._save_to_csv(file_path)
        
        if not self.workbook:
            raise ValueError("No workbook to save. Create or load a workbook first.")
        
        self.workbook.save(file_path)
        self.current_file = file_path
        _logger.info(f"Saved workbook to {file_path}")
    
    def _load_from_csv(self, file_path: str) -> Dict[str, Any]:
        """Load modpack configuration from CSV file."""
        self.current_file = file_path
        configurations = {}
        
        try:
            with open(file_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                # CSV format: modpack_name, addon_name, version, min_version, max_version, status, notes
                for row in reader:
                    modpack_name = row.get('modpack_name', '')
                    if not modpack_name:
                        continue
                    
                    if modpack_name not in configurations:
                        configurations[modpack_name] = {
                            "name": modpack_name,
                            "min_version": row.get('min_version', '1.21.0'),
                            "max_version": row.get('max_version', '1.21.90'),
                            "addons": []
                        }
                    
                    configurations[modpack_name]["addons"].append({
                        "name": row.get('addon_name', ''),
                        "path": row.get('path', ''),
                        "version": row.get('version', ''),
                        "min_version": row.get('min_version', ''),
                        "max_version": row.get('max_version', ''),
                        "status": row.get('status', 'Unknown'),
                        "notes": row.get('notes', '')
                    })
            
            _logger.info(f"Loaded {len(configurations)} modpack configurations from CSV {file_path}")
            return configurations
        except Exception as e:
            _logger.error(f"Failed to load CSV file: {e}")
            raise
    
    def _save_to_csv(self, file_path: str):
        """Save modpack configuration to CSV file."""
        if not self.csv_data:
            _logger.warning("No data to save to CSV")
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                # Write header
                writer.writerow(['modpack_name', 'addon_name', 'path', 'version', 'min_version', 'max_version', 'status', 'notes'])
                
                # Write data for each modpack
                for modpack_name, config in self.csv_data.items():
                    for addon in config['addons']:
                        writer.writerow([
                            modpack_name,
                            addon.get('name', ''),
                            addon.get('path', ''),
                            addon.get('version', ''),
                            addon.get('min_version', ''),
                            addon.get('max_version', ''),
                            addon.get('status', ''),
                            addon.get('notes', '')
                        ])
            
            self.current_file = file_path
            _logger.info(f"Saved CSV to {file_path}")
        except Exception as e:
            _logger.error(f"Failed to save CSV file: {e}")
            raise
    
    def export_modpack_to_excel(
        self,
        modpack_name: str,
        source_packs: List[str],
        output_dir: str,
        min_version: str = "1.21.0",
        max_version: str = "1.21.90"
    ):
        """
        Export a modpack configuration to Excel based on merged packs.
        
        Args:
            modpack_name: Name of the modpack
            source_packs: List of source pack file paths
            output_dir: Directory to save Excel file
            min_version: Minimum Minecraft version
            max_version: Maximum Minecraft version
        """
        if not self.workbook:
            self.create_new_workbook()
        
        # Extract addon info from pack paths
        addons = []
        for pack_path in source_packs:
            pack_name = Path(pack_path).stem
            addons.append({
                "name": pack_name,
                "path": pack_path,
                "version": "Unknown",
                "min_version": min_version,
                "max_version": max_version,
                "status": "Active",
                "notes": ""
            })
        
        self.add_modpack_sheet(modpack_name, addons, min_version, max_version)
        
        # Save to output directory
        output_path = Path(output_dir) / f"{modpack_name}_config.xlsx"
        self.save_workbook(str(output_path))
        
        _logger.info(f"Exported modpack '{modpack_name}' to {output_path}")
        return str(output_path)
    
    def import_modpack_from_excel(self, file_path: str, modpack_name: str) -> Dict[str, Any]:
        """
        Import a modpack configuration from Excel.
        
        Args:
            file_path: Path to Excel file
            modpack_name: Name of the modpack sheet to import
            
        Returns:
            Modpack configuration dictionary
        """
        configurations = self.load_from_excel(file_path)
        
        sheet_name = self._sanitize_sheet_name(modpack_name)
        if sheet_name not in configurations:
            raise ValueError(f"Modpack '{modpack_name}' not found in Excel file")
        
        return configurations[sheet_name]
    
    def get_available_modpacks(self) -> List[str]:
        """Get list of available modpack names from the workbook/CSV."""
        if self.csv_mode:
            return list(self.csv_data.keys())
        
        if not self.workbook:
            return []
        
        return [name for name in self.workbook.sheetnames if name != "Modpack Summary"]
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.current_file = None


def is_excel_available() -> bool:
    """Check if Excel/CSV functionality is available."""
    return True  # Always available (CSV mode works without openpyxl)
